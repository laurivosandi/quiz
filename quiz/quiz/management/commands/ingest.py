import ldap
import textwrap
import base64
from openpyxl import load_workbook
from datetime import datetime, timedelta
from django.core.management.base import BaseCommand, CommandError
from time import sleep
from cryptography import x509
from cryptography.hazmat.backends import default_backend
from cryptography.x509.oid import ExtensionOID, NameOID
from cryptography.x509.general_name import RFC822Name
from django.contrib.auth.models import User
from quiz.quiz.models import QuizToken, Quiz

class Command(BaseCommand):
    help = "Import Excel spreadsheet from SAIS"

    def add_arguments(self, parser):
        parser.add_argument('filename', nargs='+', type=unicode)

    def handle(self, *args, **options):
        self.conn = ldap.initialize("ldap://ldap.sk.ee")
        def generator():
            for arg in options.get("filename"):
                wb = load_workbook(filename=arg, read_only=True)
                ws = wb.active
                counter = 7
                while True:
                    try:
                        code = ws.cell(column=1, row=counter).value
                    except IndexError:
                        self.stderr.write(u"Finished reading at A%d" % counter)
                        break
                    else:
                        if not code: break
                        dt = datetime.strptime(
                            "%s %s" % (ws.cell(column=5, row=counter).value, ws.cell(column=6, row=counter).value),
                            "%d.%m.%Y %H:%M" ) - timedelta(hours=3)
                        yield code,  dt
                    counter += 1
        self.handle_list(generator())
        self.conn.unbind_s()


    def handle_list(self, m):
        group = []
        for j in m:
            group.append(j)
            if len(group) == 50:
                self.handle_group(group)
                sleep(1)
                group = []
        if group:
            self.handle_group(group)


    def handle_group(self, g):
        mapping = dict()
        codes = [j[0] for j in g]
        query = "(|" + "".join(["(serialNumber=%s)" % j[0] for j in g]) + ")"
        args = "ou=Authentication,o=ESTEID,c=EE", ldap.SCOPE_SUBTREE, query, ["serialNumber", "userCertificate;binary"]

        for dn, attributes in self.conn.search_s(*args):
            chunks = ["-----BEGIN CERTIFICATE-----"]
            for line in textwrap.wrap(base64.b64encode(attributes["userCertificate;binary"].pop()), 64):
                chunks.append(line)
            chunks.append("-----END CERTIFICATE-----")

            crt = x509.load_pem_x509_certificate("\n".join(chunks), backend=default_backend())

            for name in crt.subject:
                if name.oid == NameOID.GIVEN_NAME:
                    gn = name.value
                elif name.oid == NameOID.SURNAME:
                    sn = name.value
                elif name.oid == NameOID.SERIAL_NUMBER:
                    serial = name.value
            try:
                user = User.objects.get(username=serial)
                self.stdout.write(u"User already exists: %s" % user)
            except User.DoesNotExist:
                for extension in crt.extensions:
                    if extension.oid == ExtensionOID.SUBJECT_ALTERNATIVE_NAME:
                        for name in extension.value:
                            if isinstance(name, RFC822Name):
                                email = name.value
                user = User.objects.create(
                    email = email,
                    first_name = gn,
                    last_name = sn,
                    username = serial)
                self.stdout.write(u"Added user: %s" % user)
            mapping[serial] = user

        for serial, dt in g:
            earliest = dt - timedelta(minutes=5)
            latest = dt + timedelta(minutes=45)

            try:
                try:
                    token = QuizToken.objects.get(
                        valid = earliest,
                        user = mapping[serial])
                    self.stdout.write(u"Token already exists: %s" % token)
                except QuizToken.DoesNotExist:
                    token = QuizToken.objects.create(
                        valid = earliest,
                        expires = latest,
                        user = mapping[serial],
                        quiz = Quiz.objects.order_by('?').first(),
                        reusable = False)
                    self.stdout.write(u"Created token: %s" % token)
            except KeyError:
                self.stderr.write(u"Failed to look up: %s" % serial)
